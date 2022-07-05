import openpyxl

#список слов, которые надо игнорировать
invalid_strings_list = ["сотрудник", "аналитики", "проджектинг", "таргет", "контекст", "аккаунтинг", "smm"]
# словарик для месяцев
months = {'январь': '01', 'февраль': '02', 'март': '03', 'апрель': '04', 'май': '05', 'июнь': '06', 'июль': '07',
          'август': '08', 'сентябрь': '09', 'октябрь': '10', 'ноябрь': '11', 'декабрь': '12'}


# валидация для пустых строк и необработанных формул
def validate(unv_string):
    lower_str = str(unv_string).lower()
    return lower_str != 'none' and lower_str.lower() not in invalid_strings_list and lower_str.find('=') == -1


# допилить, если нужно
def validate_sheet(unv_sheet_name):
    return True

# функция для словарика с месяцами
def get_month(unv_month):
    reg_month = (str(unv_month).partition(' ')[0]).lower()
    return months.get(reg_month, 'XX')


# притянуть из интернета или укзаать свой путь тута
path = "D:/Twiga/Эффективность сотрудников.xlsx"

try:
    wb_obj = openpyxl.load_workbook(path, True)
except Exception:
    pass

sheet_obj = wb_obj.active
sheets = wb_obj.sheetnames

# перебираем все листы
for s in range(len(sheets)):
    sheet_obj = wb_obj[sheets[s]]
    if validate_sheet(sheets[s]):  # если какие-то листы не надо или они с другой структурой

        # тут записываю в файлики, но можно сохранить куда хотите
        with open(f'{sheets[s].capitalize()}.txt', 'w') as f:

            # перебираем всех сотрудников
            for i in range(1, sheet_obj.max_row + 1):
                # for i in range(1, 200): почему так? Многие клетки считаются активными, но там пусто, и на 20-30
                # человек он может перебирать файл до тысячной строки, что очень долго
                cell_obj = sheet_obj.cell(row=i, column=1)
                name = cell_obj.value

                if validate(name):
                    # перебираем все компании для каждого сотрудника
                    for j in range(2, sheet_obj.max_column + 1):
                        cell_obj_perc = sheet_obj.cell(row=i, column=j)
                        perc = cell_obj_perc.value
                        if validate(perc):
                            company_name = sheet_obj.cell(row=1, column=j)
                            f.write(f'{name} {get_month(sheets[s])} {company_name.value} {perc * 100}%\n')
                            # print(f'{name} {get_month(sheets[s])} {company_name.value} {perc * 100}%') для отладки на
                            # консольку
